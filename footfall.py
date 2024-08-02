import cv2
from ultralytics import YOLO
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os

# Load the YOLOv8 model
model = YOLO('yolo_models/yolov8m.pt')

# Open the video file or camera stream
video_path = 0  # 0 for webcam, or provide the video file path
cap = cv2.VideoCapture(video_path)

# Camera parameters (example values, you need to calibrate your camera to get actual values)
KNOWN_WIDTH = 0.2  # Known width of the object in meters (e.g., 20 cm)
FOCAL_LENGTH = 800  # Focal length of the camera in pixels (this is an example value)

# Video writer for saving the stream
fourcc = cv2.VideoWriter_fourcc(*'XVID')
output_filename = f"footfall_{datetime.now().strftime('%Y%m%d')}.avi"
out = cv2.VideoWriter(output_filename, fourcc, 20.0, (int(cap.get(3)), int(cap.get(4))))

# Excel file setup
excel_filename = 'footfall_data.xlsx'
if os.path.exists(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Footfall'])

def calculate_distance(known_width, focal_length, pixel_width):
    return (known_width * focal_length) / pixel_width

# Function to update footfall count in Excel
def update_footfall(date, count):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == date:
            ws.cell(row=row[0], column=2, value=row[1] + count)
            break
    else:
        ws.append([date, count])
    wb.save(excel_filename)

footfall_count = 0

# Loop through the video frames
while cap.isOpened():
    # Read a frame from the video
    success, frame = cap.read()
    frame_copy = frame.copy()
    if success:
        # Run YOLOv8 tracking on the frame, persisting tracks between frames
        results = model.track(frame, persist=True, classes=[0])  # Assuming class 0 is the object of interest

        for result in results:
            for detection in result.boxes:
                # Get the bounding box coordinates
                x1, y1, x2, y2 = map(int, detection.xyxy[0])

                # Calculate the width of the bounding box in pixels
                box_width = x2 - x1

                # Calculate the distance to the object
                distance = calculate_distance(KNOWN_WIDTH, FOCAL_LENGTH, box_width)

                # Annotate the frame with distance information on the top-right corner of the bounding box
                cv2.putText(frame_copy, f"Distance: {distance:.2f} m", (x2, y1 - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # Update footfall count
        footfall_count += len(results[0].boxes)

        # Visualize the results on the frame
        annotated_frame = results[0].plot()

        # Write the frame to the video file
        out.write(annotated_frame)

        # Display the annotated frame
        cv2.imshow("YOLOv8 Tracking", annotated_frame)

        # Break the loop if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    else:
        # Break the loop if the end of the video is reached
        break

# Update footfall data in Excel
update_footfall(datetime.now().strftime('%Y-%m-%d'), footfall_count)

# Release the video capture and writer objects, and close the display window
cap.release()
out.release()
cv2.destroyAllWindows()
